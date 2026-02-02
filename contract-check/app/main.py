"""FastAPI 应用主模块 - 遵循 LangChain 最佳实践"""

from __future__ import annotations

import asyncio
import io
import os
from contextlib import asynccontextmanager
from typing import List
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from langchain_core.runnables import Runnable
from openpyxl import load_workbook

from app.config import load_environment
from app.llm import build_checker_chain
from app.regulations import get_regulation, list_check_points, load_regulations
from app.schemas import ContractCheckResponse

# 应用级链实例 - 遵循最佳实践：链应该复用而不是每次请求都创建
_checker_chain: Runnable | None = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    应用生命周期管理。
    在应用启动时加载配置并初始化链，在关闭时清理资源。
    """
    # 启动时：加载环境变量并初始化链
    global _checker_chain
    load_environment()
    # 预加载法规配置到内存
    load_regulations()
    _checker_chain = build_checker_chain()
    yield
    # 关闭时：清理资源（如果需要）


app = FastAPI(
    title="Contract Check API",
    version="0.1.0",
    lifespan=lifespan,
)


@app.get("/v1/check-points", response_model=List[str])
async def get_check_points():
    """
    获取所有可用的审查要点列表。
    
    Returns:
        审查要点名称列表
    """
    return list_check_points()


async def _do_contract_check(check_point: str, contract: str) -> ContractCheckResponse:
    """
    内部合规检查核心逻辑，供单条检查和批量检查复用。
    
    Args:
        check_point: 审查要点
        contract: 合同内容
        
    Returns:
        ContractCheckResponse 对象
        
    Raises:
        ValueError: 参数为空或审查要点不存在
        RuntimeError: 链未初始化或调用失败
    """
    check_point_text = check_point.strip()
    contract_text = contract.strip()

    if not check_point_text:
        raise ValueError("check_point 不能为空")
    if not contract_text:
        raise ValueError("contract 不能为空")

    # 根据审查要点获取对应的法规要求
    regulation_text = get_regulation(check_point_text)  # 会抛出 KeyError

    # 使用应用级链实例（在启动时已初始化）
    if _checker_chain is None:
        raise RuntimeError("Chain not initialized. Please check application startup.")

    # 使用 LangChain 标准接口调用链
    # chain.invoke 是同步调用，使用 run_in_executor 避免阻塞事件循环
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: _checker_chain.invoke(
            {
                "regulation": regulation_text,
                "contract": contract_text,
            }
        )
    )
    return result


@app.post(
    "/v1/contract-check",
    response_model=ContractCheckResponse,
    response_model_by_alias=True,  # 输出中文字段：审查结果/审查过程/审查建议
)
async def contract_check(
    check_point: str = Form(..., description="审查要点（如：主体资格审查）"),
    contract: str = Form(..., description="合同内容文本"),
):
    try:
        return await _do_contract_check(check_point, contract)
    except KeyError as e:
        available_points = list_check_points()
        raise HTTPException(
            status_code=400,
            detail=f"未知的审查要点: '{check_point}'。可用的审查要点: {available_points}",
        ) from e
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"LLM contract-check failed: {type(e).__name__}: {e}",
        ) from e


@app.post("/v1/contract-check-batch")
async def contract_check_batch(
    file: UploadFile = File(..., description="Excel 文件，包含「审查项」和「合同原文」列"),
    concurrency: int = Form(default=5, ge=1, le=20, description="并发数量（1-20）"),
):
    """
    批量处理 Excel 文件中的合同审查。
    
    Excel 文件要求：
    - 第一行为表头
    - 必须包含「审查项」列（对应 regulations.yaml 中的 key）
    - 必须包含「合同原文」列（需要审查的合同内容）
    - 结果会写入「审查结果」「审查过程」「审查建议」列
    
    Args:
        file: 上传的 Excel 文件（.xlsx 格式）
        concurrency: 并发处理数量，默认 5
        
    Returns:
        处理后的 Excel 文件（StreamingResponse）
    """
    # 验证文件类型
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="请上传 Excel 文件（.xlsx 或 .xls 格式）"
        )
    
    # 读取 Excel 文件
    try:
        content = await file.read()
        workbook = load_workbook(filename=io.BytesIO(content))
        sheet = workbook.worksheets[0]
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"无法读取 Excel 文件: {str(e)}"
        ) from e
    
    # 解析表头，找到关键列的索引
    headers = [cell.value.strip() for cell in sheet[1]]   # openpyxl 行索引从 1 开始
    
    try:
        check_point_col = headers.index("审查项") + 1  # openpyxl 列索引从 1 开始
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Excel 文件缺少「审查项」列"
        )
    
    try:
        contract_col = headers.index("合同原文") + 1
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Excel 文件缺少「合同原文」列"
        )
    
    # 查找或创建结果列（result_col 带模型名，便于区分不同模型的输出）
    model_type = (os.getenv("MODEL_TYPE") or "auto").strip().lower()
    result_col_name = f"{model_type}审查结果"
    process_col_name = f"{model_type}审查过程"
    suggestion_col_name = f"{model_type}审查建议"
    
    def get_or_create_col(col_name: str) -> int:
        """获取列索引，如果不存在则创建新列"""
        if col_name in headers:
            return headers.index(col_name) + 1
        else:
            new_col = len(headers) + 1
            sheet.cell(row=1, column=new_col, value=col_name)
            headers.append(col_name)
            return new_col
    
    result_col = get_or_create_col(result_col_name)
    process_col = get_or_create_col(process_col_name)
    suggestion_col = get_or_create_col(suggestion_col_name)
    
    # 收集需要处理的行
    rows_to_process = []
    for row_idx in range(2, sheet.max_row + 1):  # 从第 2 行开始（跳过表头）
        check_point = sheet.cell(row=row_idx, column=check_point_col).value
        contract = sheet.cell(row=row_idx, column=contract_col).value
        
        if check_point and contract:
            rows_to_process.append((row_idx, str(check_point), str(contract)))
    
    if not rows_to_process:
        raise HTTPException(
            status_code=400,
            detail="Excel 文件中没有有效的数据行（审查项和合同原文都不能为空）"
        )
    
    # 使用信号量控制并发数
    semaphore = asyncio.Semaphore(concurrency)
    
    async def process_row(row_idx: int, check_point: str, contract: str) -> tuple[int, str, str, str]:
        """处理单行数据"""
        async with semaphore:
            try:
                result = await _do_contract_check(check_point, contract)
                return (
                    row_idx,
                    result.review_result.value,  # 审查结果: "合格" 或 "不合格"
                    result.review_process,       # 审查过程
                    result.review_suggestion,    # 审查建议
                )
            except KeyError:
                return (row_idx, "错误", f"未知的审查要点: {check_point}", "请检查审查项是否正确")
            except Exception as e:
                return (row_idx, "错误", f"处理失败: {type(e).__name__}: {str(e)}", "请重试或检查输入")
    
    # 并发处理所有行
    tasks = [process_row(row_idx, cp, ct) for row_idx, cp, ct in rows_to_process]
    results = await asyncio.gather(*tasks)
    
    # 将结果写入 Excel
    for row_idx, review_result, review_process, review_suggestion in results:
        sheet.cell(row=row_idx, column=result_col, value=review_result)
        sheet.cell(row=row_idx, column=process_col, value=review_process)
        sheet.cell(row=row_idx, column=suggestion_col, value=review_suggestion)
    
    # 将处理后的 Excel 保存到内存
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # 生成输出文件名
    original_name = file.filename or "result"
    if original_name.endswith('.xlsx'):
        output_name = original_name[:-5] + "_审查结果.xlsx"
    elif original_name.endswith('.xls'):
        output_name = original_name[:-4] + "_审查结果.xlsx"
    else:
        output_name = original_name + "_审查结果.xlsx"
    
    # 使用 URL 编码处理文件名，避免 latin-1 编码错误
    # RFC 5987 格式: filename*=UTF-8''encoded_filename
    encoded_filename = quote(output_name, safe='')
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@app.get("/healthz")
def healthz():
    return JSONResponse({"ok": True})


if __name__ == "__main__":
    import os
    import sys

    # 在导入 uvicorn 之前禁用 uvloop
    os.environ["UVLOOP"] = "false"

    # 检查是否在调试模式下运行
    if "pydevd" in sys.modules:
        print("Running in debug mode, forcing asyncio loop")

    import uvicorn

    uvicorn.run(
        "app.main:app",
        host="127.0.0.1",
        port=8000,
        reload=False,
        workers=1,
        loop="asyncio"  # 显式指定使用 asyncio 事件循环
    )
